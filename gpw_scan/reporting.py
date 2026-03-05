# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import json
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def df_to_clickable_html(df: pd.DataFrame, n: int = 30) -> str:
    if df is None or df.empty:
        return "<p><em>Brak pozycji.</em></p>"

    dff = df.head(n).copy()
    if "Symbol" in dff.columns:
        dff["Symbol"] = dff["Symbol"].astype(str).apply(
            lambda s: f'<a href="#" class="sym" data-sym="{s}">{s}</a>'
        )

    html_table = dff.to_html(index=False, escape=False)
    return f'<div class="tableWrap">{html_table}</div>'


def build_interactive_html(
    cfg: dict,
    df_leaders: pd.DataFrame,
    df_breakouts_watch: pd.DataFrame,
    df_breakouts_strict: pd.DataFrame,
    df_pullbacks: pd.DataFrame,
    interactive_store: dict,
) -> str:
    topN = int(cfg["TOP_N_TABLES"])
    data_json = json.dumps(interactive_store, ensure_ascii=False).replace("</", "<\\/")

    js_cfg = {
        "GAUGE_RED_LT": cfg["GAUGE_THRESHOLDS"]["RED_LT"],
        "GAUGE_YELLOW_LE": cfg["GAUGE_THRESHOLDS"]["YELLOW_LE"],
        "ADX_REF_LINE": cfg["ADX_REF_LINE"],
        "EMA_FAST_D": cfg["EMA_FAST_D"],
        "EMA_SLOW_D": cfg["EMA_SLOW_D"],
    }

    interactive_html = f"""
<html><head><meta charset="utf-8">
<title>GPW_SCAN v19b — Swing (D1)</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>

<style>
:root {{
  --bg: #0b0f17;
  --panel: #111827;
  --panel2: #0f172a;
  --text: #e5e7eb;
  --muted: #9ca3af;
  --border: rgba(255,255,255,0.08);
  --accent: #38bdf8;
  --red: #fb7185;
  --yellow: #fbbf24;
  --green: #34d399;
}}
body {{ margin:0; font-family: Arial, sans-serif; background: var(--bg); color: var(--text); }}
header {{ padding: 14px 18px; border-bottom: 1px solid var(--border);
  background: linear-gradient(180deg, rgba(56,189,248,0.10), rgba(0,0,0,0)); }}

.container {{ display: grid; grid-template-columns: 230px 1fr; height: calc(100vh - 62px); }}
.left {{ overflow: auto; padding: 10px; border-right: 1px solid var(--border); }}
.right {{ overflow: auto; padding: 14px; }}

.card {{ border: 1px solid var(--border); border-radius: 16px; padding: 10px; margin: 10px 0;
  background: var(--panel); box-shadow: 0 10px 30px rgba(0,0,0,0.25); }}
h2 {{ margin: 6px 0 10px 0; font-size: 14px; letter-spacing: .2px; }}

a.sym {{ font-weight: 800; color: var(--accent); text-decoration: none; }}
a.sym:hover {{ text-decoration: underline; }}

.search {{ width: 100%; padding: 10px; font-size: 13px; border: 1px solid var(--border);
  border-radius: 12px; background: var(--panel2); color: var(--text); outline: none; }}
.small {{ font-size: 12px; color: var(--muted); }}

.chart {{ height: 320px; }}
.chartTall {{ height: 420px; }}

.tableWrap {{
  width: 100%;
  overflow-x: auto;
  overflow-y: hidden;
  border-radius: 12px;
  border: 1px solid var(--border);
}}
.tableWrap table {{
  border-collapse: collapse;
  table-layout: auto;
  width: max-content;
  min-width: 980px;
}}
.tableWrap th, .tableWrap td {{
  border: 1px solid var(--border);
  padding: 6px 8px;
  font-size: 12px;
  white-space: nowrap;
  color: var(--text);
}}
.tableWrap th {{
  background: rgba(255,255,255,0.04);
  position: static !important;
}}

.kpis {{
  display:grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 10px;
}}
.kpi {{
  border: 1px solid var(--border); border-radius: 14px; padding: 10px;
  background: rgba(255,255,255,0.03);
}}
.kpi .lab {{ font-size: 11px; color: var(--muted); }}
.kpi .val {{ font-size: 16px; font-weight: 900; margin-top: 4px; }}
.kpi .hint {{ font-size: 11px; color: var(--muted); margin-top: 4px; }}

.gauges {{
  display:grid;
  grid-template-columns: 1fr 1fr 1fr;
  gap: 12px;
  margin-top: 10px;
}}
.gbox {{
  border: 1px solid var(--border);
  border-radius: 14px;
  padding: 8px;
  background: rgba(255,255,255,0.03);
}}
.gtitle {{
  font-size: 12px;
  color: var(--muted);
  margin: 2px 0 6px 2px;
  display:flex;
  justify-content:space-between;
  align-items:center;
  gap: 8px;
}}
.gauge {{ height: 180px; }}
.badge {{
  padding: 4px 8px;
  border: 1px solid var(--border);
  border-radius: 999px;
  font-size: 11px;
  color: var(--text);
  background: rgba(255,255,255,0.03);
}}
.toggle {{ display:flex; gap:8px; align-items:center; flex-wrap:wrap; }}
.btn {{
  cursor:pointer; padding: 6px 10px; border-radius: 999px;
  border: 1px solid var(--border); background: rgba(255,255,255,0.03);
  color: var(--text); font-size: 12px;
}}
.btn.active {{ border-color: rgba(56,189,248,0.6); box-shadow: 0 0 0 3px rgba(56,189,248,0.12); }}
.hr {{ height:1px; background: var(--border); margin: 10px 0; }}
</style>
</head><body>

<header>
  <div style="display:flex; align-items:center; justify-content:space-between; gap:12px;">
    <div>
      <div style="font-size:16px; font-weight:900;">GPW_SCAN v19b — Swing (D1)</div>
      <div class="small">Kliknij symbol po lewej → KPI + GAUGE + wykresy. Scoring jest na D1 (daily close).</div>
    </div>
  </div>
</header>

<div class="container">
  <div class="left">
    <input id="search" class="search" placeholder="Szukaj symbolu..." />

    <div class="card">
      <h2>Leaders</h2>
      {df_to_clickable_html(df_leaders, topN)}
      <div class="small">Warunek: Liquid_OK=True oraz Score10 ≥ {cfg["LEADERS_SCORE10_MIN"]}.</div>
    </div>

    <div class="card">
      <h2>Breakouts — Watchlist</h2>
      {df_to_clickable_html(df_breakouts_watch, topN)}
      <div class="small">Warunek: IsBreakout=True oraz Score10 ≥ {cfg["BREAKOUTS_WATCH_SCORE10_MIN"]}.</div>
    </div>

    <div class="card">
      <h2>Breakouts — Strict</h2>
      {df_to_clickable_html(df_breakouts_strict, topN)}
      <div class="small">Warunek: breakout + Liquid_OK + MACD(D)>Signal + VolSpike ≥ {cfg["VOL_SPIKE_MULT_STRICT"]}.</div>
    </div>

    <div class="card">
      <h2>Pullbacks</h2>
      {df_to_clickable_html(df_pullbacks, topN)}
      <div class="small">Warunek: Liquid_OK + Score10≥{cfg["PULLBACKS_SCORE10_MIN"]} i blisko EMA.</div>
    </div>
  </div>

  <div class="right">
    <div class="card" id="card_top">
      <div style="display:flex; align-items:flex-start; justify-content:space-between; gap:12px;">
        <div>
          <div id="title" style="font-size:18px; font-weight:1000;">Wybierz instrument</div>
          <div id="subtitle" class="small">Kliknij symbol w tabeli po lewej stronie.</div>
        </div>

        <div class="toggle">
          <span class="small">Daily window:</span>
          <button id="btnD10" class="btn">10</button>
          <button id="btnD20" class="btn active">20</button>
          <button id="btnD40" class="btn">40</button>
          <button id="btnD60" class="btn">60</button>
          <button id="btnD120" class="btn">120</button>

          <span class="small" style="margin-left:10px;">Weekly (pogląd):</span>
          <button id="btnW30" class="btn active">30W</button>
          <button id="btnW52" class="btn">52W</button>
        </div>
      </div>

      <div class="hr"></div>

      <div id="kpis" class="kpis"></div>

      <div class="gauges">
        <div class="gbox">
          <div class="gtitle"><span>TQ — Trend Quality</span><span id="tq_lab" class="badge">—</span></div>
          <div id="g_tq" class="gauge"></div>
        </div>
        <div class="gbox">
          <div class="gtitle"><span>PQ — Participation</span><span id="pq_lab" class="badge">—</span></div>
          <div id="g_pq" class="gauge"></div>
        </div>
        <div class="gbox">
          <div class="gtitle"><span>RQ — Risk & Smoothness</span><span id="rq_lab" class="badge">—</span></div>
          <div id="g_rq" class="gauge"></div>
        </div>
      </div>
    </div>

    <!-- ✅ NOWY UKŁAD: wszystkie dzienne wykresy full width, w kolejności jak podałeś -->

    <div class="card">
      <h2>Daily Candles + EMA (scoring: D1)</h2>
      <div id="chart_daily_candles" class="chartTall"></div>
    </div>

    <div class="card">
      <h2>Daily Volume</h2>
      <div id="chart_vol" class="chartTall"></div>
    </div>

    <div class="card">
      <h2>Daily MACD</h2>
      <div id="chart_macd" class="chartTall"></div>
    </div>

    <div class="card">
      <h2>Daily Donchian</h2>
      <div id="chart_don" class="chartTall"></div>
    </div>

    <div class="card">
      <h2>Daily ADX</h2>
      <div id="chart_adx" class="chartTall"></div>
    </div>

    <div class="card">
      <h2>Daily RSI (Wilder)</h2>
      <div id="chart_rsi" class="chartTall"></div>
    </div>

    <div class="card">
      <h2>Weekly Candles (pogląd)</h2>
      <div id="chart_weekly" class="chartTall"></div>
      <div class="small">Weekly jest tylko do kontekstu na wykresie, nie wpływa na scoring.</div>
    </div>

  </div>
</div>

<script>
const DATA = {data_json};
const CFG = {json.dumps(js_cfg)};
let CURRENT_SYM = null;
let WINDOW_D = 20;
let WINDOW_W = 30;

// ---------- format helpers ----------
function fmtPct(x){{
  if(x===null || x===undefined || isNaN(x)) return "—";
  return (x*100).toFixed(2) + "%";
}}
function fmtNum(x, digits=2){{
  if(x===null || x===undefined || isNaN(x)) return "—";
  return Number(x).toFixed(digits);
}}
function fmtPLN(x){{
  if(x===null || x===undefined || isNaN(x)) return "—";
  try {{
    return new Intl.NumberFormat('pl-PL', {{ maximumFractionDigits: 0 }}).format(x) + " PLN";
  }} catch(e) {{
    return Math.round(x) + " PLN";
  }}
}}
function sliceTail(arr, n){{
  if(!arr) return [];
  return arr.slice(Math.max(0, arr.length - n));
}}
function returnsFromBase(series){{
  const s = series.map(x => (x===null||x===undefined) ? null : Number(x));
  const first = s.find(v => v!==null && !isNaN(v) && v!==0);
  if(first===undefined || first===null || first===0) return s.map(_=>null);
  return s.map(v => (v===null || isNaN(v)) ? null : (v/first - 1.0));
}}
function setActiveBtns(){{
  document.getElementById("btnD10").classList.toggle("active", WINDOW_D===10);
  document.getElementById("btnD20").classList.toggle("active", WINDOW_D===20);
  document.getElementById("btnD40").classList.toggle("active", WINDOW_D===40);
  document.getElementById("btnD60").classList.toggle("active", WINDOW_D===60);
  document.getElementById("btnD120").classList.toggle("active", WINDOW_D===120);
  document.getElementById("btnW30").classList.toggle("active", WINDOW_W===30);
  document.getElementById("btnW52").classList.toggle("active", WINDOW_W===52);
}}

// --- legend-under helper (Plotly) ---
function legendUnder(){{
  return {{
    orientation: "h",
    x: 0,
    y: -0.28,
    xanchor: "left",
    yanchor: "top"
  }};
}}
function baseLayout(extra={{}}){{
  return Object.assign({{
    paper_bgcolor:"rgba(0,0,0,0)",
    plot_bgcolor:"rgba(0,0,0,0)",
    font:{{color:"var(--text)"}},
    margin:{{t:10,r:10,b:80,l:50}},   // ✅ większy dół na legendę pod wykresem
  }}, extra);
}}

function gaugeColor(v){{
  if(v < CFG.GAUGE_RED_LT) return "var(--red)";
  if(v <= CFG.GAUGE_YELLOW_LE) return "var(--yellow)";
  return "var(--green)";
}}
function drawGauge(divId, value){{
  const v = (value===null || value===undefined || isNaN(value)) ? 0 : Number(value);
  const col = gaugeColor(v);

  const data = [{{
    type: "indicator",
    mode: "gauge+number",
    value: v,
    number: {{ suffix: "" }},
    title: {{ text: "" }},
    gauge: {{
      shape: "angular",
      axis: {{ range: [0, 100], tickwidth: 1, ticks: "" }},
      bar: {{ thickness: 0.35, color: col }},
      steps: [
        {{ range: [0, CFG.GAUGE_RED_LT], color: "rgba(251,113,133,0.12)" }},
        {{ range: [CFG.GAUGE_RED_LT, CFG.GAUGE_YELLOW_LE], color: "rgba(251,191,36,0.12)" }},
        {{ range: [CFG.GAUGE_YELLOW_LE, 100], color: "rgba(52,211,153,0.12)" }},
      ],
    }}
  }}];

  const layout = {{
    paper_bgcolor: "rgba(0,0,0,0)",
    plot_bgcolor: "rgba(0,0,0,0)",
    margin: {{ t: 10, r: 10, b: 10, l: 10 }},
    font: {{ color: "var(--text)" }},
  }};

  Plotly.newPlot(divId, data, layout, {{displayModeBar: false}});
}}
function kpiCard(label, value, hint){{
  return (
    '<div class="kpi">' +
      '<div class="lab">' + String(label) + '</div>' +
      '<div class="val">' + String(value) + '</div>' +
      '<div class="hint">' + String(hint || "") + '</div>' +
    '</div>'
  );
}}
function renderKPIs(p){{
  const k = (p && p.kpi) ? p.kpi : {{}};
  const r14 = fmtNum(k.RSI14_D, 1);
  const r10 = fmtNum(k.RSI10_D, 1);

  const html = [
    kpiCard("Score10", (k.Score10 ?? "—"), "0–10 (D1 trend/setup)"),
    kpiCard("Score4",  (k.Score4  ?? "—"), "0–4 (szybki filtr)"),
    kpiCard("1D Close", fmtNum(k.D_close, 2), "day: " + (k.D_last || "—")),
    kpiCard("Return 20D", fmtPct(k.Return_20D), "ok. 1 mies."),
    kpiCard("Return 60D", fmtPct(k.Return_60D), "ok. 3 mies."),
    kpiCard("RS 60D", fmtPct(k.RS_60D), "vs benchmark (jeśli dostępny)"),
    kpiCard("ADX14", fmtNum(k.ADX14,1), "≥ " + CFG.ADX_REF_LINE + " = trend"),
    kpiCard("ATR% (D)", fmtPct(k.ATR14_pct), "zmienność"),
    kpiCard("OZ factor (RSI)", (r14 + " / " + r10), "RSI14/RSI10"),
    kpiCard("VolSpike", fmtNum(k.VolSpikeRatio,2), "D_vol / SMA20"),
    kpiCard("Turnover med20", fmtPLN(k.TurnoverApproxPLN_med20), "close*vol (proxy)")
  ].join("");

  document.getElementById("kpis").innerHTML = html;

  document.getElementById("tq_lab").textContent = (k.TQ_Label || "—");
  document.getElementById("pq_lab").textContent = (k.PQ_Label || "—");
  document.getElementById("rq_lab").textContent = (k.RQ_Label || "—");

  drawGauge("g_tq", k.TQ);
  drawGauge("g_pq", k.PQ);
  drawGauge("g_rq", k.RQ);
}}

function draw(sym){{
  const p = DATA[sym];
  if(!p) return;

  CURRENT_SYM = sym;
  document.getElementById("title").textContent = sym + " — " + (p.Nazwa || "");
  document.getElementById("subtitle").textContent = "KPI → GAUGE → wykresy (scoring: D1).";

  renderKPIs(p);

  // ---- DAILY ----
  const d = p.daily || {{}};
  const xd = d.dates || [];
  const dopen = d.open || [];
  const dhigh = d.high || [];
  const dlow  = d.low  || [];
  const dclose = d.close || [];
  const dvol  = d.vol || [];

  const emaF = d.ema_fast || [];
  const emaS = d.ema_slow || [];
  const donu = d.don_u || [];
  const donl = d.don_l || [];
  const adx  = d.adx14 || [];
  const macd = d.macd || [];
  const sig  = d.macd_signal || [];
  const hist = d.macd_hist || [];
  const rsi14 = d.rsi14 || [];
  const rsi10 = d.rsi10 || [];

  const xD  = sliceTail(xd, WINDOW_D);
  const oD  = sliceTail(dopen, WINDOW_D);
  const hD  = sliceTail(dhigh, WINDOW_D);
  const lD  = sliceTail(dlow, WINDOW_D);
  const cD  = sliceTail(dclose, WINDOW_D);
  const vD  = sliceTail(dvol, WINDOW_D);

  const eF  = sliceTail(emaF, WINDOW_D);
  const eS  = sliceTail(emaS, WINDOW_D);

  // 1) Daily Candles + EMA
  Plotly.newPlot("chart_daily_candles", [
    {{ x:xD, open:oD, high:hD, low:lD, close:cD, type:"candlestick", name:"OHLC (D)" }},
    {{ x:xD, y:eF, name:"EMA"+CFG.EMA_FAST_D+"D", mode:"lines" }},
    {{ x:xD, y:eS, name:"EMA"+CFG.EMA_SLOW_D+"D", mode:"lines" }},
  ], baseLayout({{
    xaxis:{{title:"Day"}}, yaxis:{{title:"Price"}},
    legend:{{orientation:"h"}}
  }}), {{displayModeBar:false}});

  // 2) Daily Volume (full width)
  Plotly.newPlot("chart_vol", [
    {{ x: xD, y: vD, name:"Vol", type:"bar" }},
  ], baseLayout({{
    xaxis:{{title:"Day"}}, yaxis:{{title:"Volume"}},
    legend:{{orientation:"h"}}
  }}), {{displayModeBar:false}});

  // 3) Daily MACD (legend under)
  const mD = sliceTail(macd, WINDOW_D);
  const sD = sliceTail(sig, WINDOW_D);
  const hD2= sliceTail(hist, WINDOW_D);

  Plotly.newPlot("chart_macd", [
    {{ x: xD, y: mD, name:"MACD(D)", mode:"lines" }},
    {{ x: xD, y: sD, name:"Signal", mode:"lines" }},
    {{ x: xD, y: hD2, name:"Hist", type:"bar" }},
  ], baseLayout({{
    xaxis:{{title:"Day"}},
    legend: legendUnder()
  }}), {{displayModeBar:false}});

  // 4) Daily Donchian (legend under)
  Plotly.newPlot("chart_don", [
    {{ x: xd, y: dclose, name:"Close", mode:"lines" }},
    {{ x: xd, y: donu,   name:"DonchianU", mode:"lines" }},
    {{ x: xd, y: donl,   name:"DonchianL", mode:"lines" }},
  ], baseLayout({{
    xaxis:{{title:"Day"}},
    legend: legendUnder()
  }}), {{displayModeBar:false}});

  // 5) Daily ADX (legend under)
  const adxLine = xd.map(_ => CFG.ADX_REF_LINE);
  Plotly.newPlot("chart_adx", [
    {{ x: xd, y: adx, name:"ADX", mode:"lines" }},
    {{ x: xd, y: adxLine, name:"ADX="+CFG.ADX_REF_LINE, mode:"lines", line:{{dash:"dot"}} }},
  ], baseLayout({{
    xaxis:{{title:"Day"}}, yaxis:{{title:"ADX"}},
    legend: legendUnder()
  }}), {{displayModeBar:false}});

  // 6) RSI (jak było)
  const r14 = sliceTail(rsi14, WINDOW_D);
  const r10 = sliceTail(rsi10, WINDOW_D);
  const l30 = xD.map(_=>30);
  const l50 = xD.map(_=>50);
  const l70 = xD.map(_=>70);

  Plotly.newPlot("chart_rsi", [
    {{ x:xD, y:r14, name:"RSI14", mode:"lines" }},
    {{ x:xD, y:r10, name:"RSI10", mode:"lines" }},
    {{ x:xD, y:l30, name:"30", mode:"lines", line:{{dash:"dot"}} }},
    {{ x:xD, y:l50, name:"50", mode:"lines", line:{{dash:"dot"}} }},
    {{ x:xD, y:l70, name:"70", mode:"lines", line:{{dash:"dot"}} }},
  ], baseLayout({{
    xaxis:{{title:"Day"}},
    yaxis:{{title:"RSI", range:[0,100]}},
    legend:{{orientation:"h"}}
  }}), {{displayModeBar:false}});

  // 7) Weekly (jak było)
  const w = p.weekly || {{}};
  const xw = sliceTail(w.dates || [], WINDOW_W);
  const oW = sliceTail(w.open || [], WINDOW_W);
  const hW = sliceTail(w.high || [], WINDOW_W);
  const lW = sliceTail(w.low  || [], WINDOW_W);
  const cW = sliceTail(w.close|| [], WINDOW_W);

  Plotly.newPlot("chart_weekly", [
    {{ x: xw, open: oW, high: hW, low: lW, close: cW, type:"candlestick", name:"OHLC (W)" }},
  ], baseLayout({{
    xaxis:{{title:"Week"}}, yaxis:{{title:"Price"}},
    legend:{{orientation:"h"}}
  }}), {{displayModeBar:false}});
}}

function hookLinks(){{
  document.querySelectorAll("a.sym").forEach(a => {{
    a.addEventListener("click", (e) => {{
      e.preventDefault();
      draw(a.getAttribute("data-sym"));
    }});
  }});
}}
hookLinks();

setTimeout(() => {{
  const first = document.querySelector("a.sym");
  if(first) first.click();
}}, 50);

document.getElementById("btnD10").addEventListener("click", () => {{ WINDOW_D=10; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});
document.getElementById("btnD20").addEventListener("click", () => {{ WINDOW_D=20; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});
document.getElementById("btnD40").addEventListener("click", () => {{ WINDOW_D=40; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});
document.getElementById("btnD60").addEventListener("click", () => {{ WINDOW_D=60; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});
document.getElementById("btnD120").addEventListener("click", () => {{ WINDOW_D=120; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});

document.getElementById("btnW30").addEventListener("click", () => {{ WINDOW_W=30; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});
document.getElementById("btnW52").addEventListener("click", () => {{ WINDOW_W=52; setActiveBtns(); if(CURRENT_SYM) draw(CURRENT_SYM); }});

document.getElementById("search").addEventListener("input", (e) => {{
  const q = (e.target.value || "").toUpperCase().trim();
  document.querySelectorAll("a.sym").forEach(a => {{
    const sym = (a.getAttribute("data-sym")||"").toUpperCase();
    a.style.background = (q && sym.includes(q)) ? "rgba(251,191,36,0.25)" : "transparent";
    a.style.borderRadius = (q && sym.includes(q)) ? "8px" : "0px";
    a.style.padding = (q && sym.includes(q)) ? "2px 4px" : "0px";
  }});
}});

setActiveBtns();
</script>

</body></html>
"""
    return interactive_html


def write_excel_bytes(
    cfg: dict,
    df_rank: pd.DataFrame,
    df_leaders: pd.DataFrame,
    df_breakouts_watch: pd.DataFrame,
    df_breakouts_strict: pd.DataFrame,
    df_pullbacks: pd.DataFrame,
    df_err: pd.DataFrame,
) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_rank.to_excel(writer, index=False, sheet_name="Ranking")
        df_leaders.to_excel(writer, index=False, sheet_name="Leaders")
        df_breakouts_watch.to_excel(writer, index=False, sheet_name="Breakouts_Watch")
        df_breakouts_strict.to_excel(writer, index=False, sheet_name="Breakouts_Strict")
        df_pullbacks.to_excel(writer, index=False, sheet_name="Pullbacks")
        df_err.to_excel(writer, index=False, sheet_name="Errors")

    out.seek(0)
    wb = load_workbook(out)

    PLN_FORMAT = cfg["PLN_FORMAT"]
    PCT_FORMAT = cfg["PCT_FORMAT"]
    NUM_FORMAT = cfg["NUM_FORMAT"]

    CURRENCY_COLS = {
        "D_close","EMA_fast_D","EMA_slow_D","DonchianU_20","DonchianU_20_prev","DonchianL_20","TurnoverApproxPLN_med20"
    }
    PERCENT_COLS  = {
        "Return_20D","Return_60D","ATR14_pct","DistToEMA_fast_D","DistToEMA_slow_D","RS_60D","TrendPersist40D"
    }
    NUMBER_COLS   = {"TQ","PQ","RQ","ADX14","VolSpikeRatio","RSI14_D","RSI10_D"}

    fill_strong = PatternFill(
        start_color=cfg["HIGHLIGHT_RULES"]["FILL_STRONG"],
        end_color=cfg["HIGHLIGHT_RULES"]["FILL_STRONG"],
        fill_type="solid",
    )
    fill_good = PatternFill(
        start_color=cfg["HIGHLIGHT_RULES"]["FILL_GOOD"],
        end_color=cfg["HIGHLIGHT_RULES"]["FILL_GOOD"],
        fill_type="solid",
    )

    def format_sheet(ws):
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            name = ws.cell(row=1, column=col_idx).value
            if isinstance(name, str) and name.strip():
                headers[name.strip()] = col_idx

        def apply_format(col_name: str, number_format: str):
            if col_name not in headers:
                return
            col_idx = headers[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = number_format

        for c in CURRENCY_COLS:
            apply_format(c, PLN_FORMAT)
        for c in PERCENT_COLS:
            apply_format(c, PCT_FORMAT)
        for c in NUMBER_COLS:
            apply_format(c, NUM_FORMAT)

        score10_col = headers.get("Score10")
        score4_col  = headers.get("Score4")
        if score10_col and score4_col:
            strong_s10 = set(cfg["HIGHLIGHT_RULES"]["STRONG"]["Score10"])
            strong_s4  = set(cfg["HIGHLIGHT_RULES"]["STRONG"]["Score4"])
            good_s10   = set(cfg["HIGHLIGHT_RULES"]["GOOD"]["Score10"])
            good_s4    = set(cfg["HIGHLIGHT_RULES"]["GOOD"]["Score4"])

            for r in range(2, ws.max_row + 1):
                v10 = ws.cell(row=r, column=score10_col).value
                v4  = ws.cell(row=r, column=score4_col).value
                try:
                    v10n = int(v10) if v10 is not None else None
                except Exception:
                    v10n = None
                try:
                    v4n = int(v4) if v4 is not None else None
                except Exception:
                    v4n = None

                fill = None
                if (v10n in strong_s10) and (v4n in strong_s4):
                    fill = fill_strong
                elif (v10n in good_s10) and (v4n in good_s4):
                    fill = fill_good

                if fill is not None:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for rr in range(1, ws.max_row + 1):
                v = ws.cell(row=rr, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

    for sh in ["Ranking", "Leaders", "Breakouts_Watch", "Breakouts_Strict", "Pullbacks"]:
        if sh in wb.sheetnames:
            format_sheet(wb[sh])

    out2 = io.BytesIO()
    wb.save(out2)
    return out2.getvalue()